# -*- coding: utf-8 -*-
"""
Created on Tue Sep  7 20:04:30 2021

@author: jan_c
"""
#%% nombres de los archivos de excel y los complejos que se acomodará
import pandas as pd
from openpyxl import load_workbook

archivos = ['WACFHZ-libre_DMSO-Agua_reanalisis',
            'WACFHZ-NaCH3CO2_DMSO-Agua_reanalisis',
            'WACFHZ-LiCl_DMSO-Agua_reanalisis',
            'WACFHZ-Na2SO4_DMSO-Agua_reanalisis',
            'WACFHZ-NaF_DMSO-Agua_reanalisis',
            'WACFHZ-Na2HPO4_DMSO-Agua_reanalisis',
            'WACFHZ-KI_DMSO-Agua_reanalisis',
            'WACFHZ-KCl_DMSO-Agua_reanalisis',
            'WACFHZ-NaCl_DMSO-Agua_reanalisis']

receptor = ('WACFHZ',
            'WACFHZ-NaCOO',
            'WACFHZ-LiCl',
            'WACFHZ-Na2SO4',
            'WACFHZ-NaF',
            'WACFHZ-Na2HPO4',
            'WACFHZ-KI',
            'WACFHZ-KCl',
            'WACFHZ-NaCl')

complejos = ('WACFHZ',
            'WACFHZ_NaCOO',
            'WACFHZ_LiCl',
            'WACFHZ_Na2SO4',
            'WACFHZ_NaF',
            'WACFHZ_Na2HPO4',
            'WACFHZ_KI',
            'WACFHZ_KCl',
            'WACFHZ_NaCl')
            
#%% función para extraer la información importante
def acomodo_datos(nombre_libro, receptor):
    #Leer archivo de entrada
    libro = nombre_libro
    libro_datos = libro + ".xlsx" # Asigna el nombre correcto del archivo a leer
    datos = pd.read_excel(libro_datos, sheet_name="Hoja1")

    # Se filtran las columnas de interes y se generan los datos ordenados
    filtro_fluorescencia = datos.filter(regex = "Fluorescencia") # filtra columnas
    datos_f = pd.DataFrame(filtro_fluorescencia)
    muestra = pd.DataFrame(datos["Muestra"])
    datos_gen = pd.concat([muestra, datos_f], axis=1) # Concatenamos los datos
    datos_generales = datos_gen.set_index("Muestra")
    datos_i = datos_generales.loc[receptor]
    print(datos_i)
    return datos_i

#%% Se crea el archivo de excel vacio
data_null = {}
df_null = pd.DataFrame(data_null)
df_null.to_excel("Datos_de_salida.xlsx", index = False)

#%% Se crea el ciclo para cargar las hojas en excel

for i, j in zip(archivos, receptor):
    data = acomodo_datos(i, j) / 113.4 #se divide entre el valor promedio de R libre.
    book = load_workbook("Datos_de_salida.xlsx")
    with pd.ExcelWriter("Datos_de_salida.xlsx") as writer:
        writer.book = book 
        data.to_excel(writer, sheet_name = j)

#%% Añadire un libro con todos los datos encimados

for i, j in zip(receptor, complejos): #tuve que generar un diccionario para nombres (comlejos)
    ac_datos = pd.read_excel("Datos_de_salida.xlsx", sheet_name=i)
    globals()['c_%s' % j] = ac_datos #genera nombres en el ciclo for

datos_totales = pd.concat([c_WACFHZ, c_WACFHZ_NaCOO, c_WACFHZ_LiCl,
            c_WACFHZ_Na2SO4, c_WACFHZ_NaF, c_WACFHZ_Na2HPO4, c_WACFHZ_KI,
            c_WACFHZ_KCl, c_WACFHZ_NaCl], axis = 0)

datos_totales = datos_totales.set_index("Muestra")

book = load_workbook("Datos_de_salida.xlsx")
with pd.ExcelWriter("Datos_de_salida.xlsx") as writer:
    writer.book = book 
    datos_totales.to_excel(writer, sheet_name = "datos_totales")