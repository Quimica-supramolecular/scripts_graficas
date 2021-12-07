# -*- coding: utf-8 -*-
"""
Created on Mon Oct 11 11:11:03 2021

@author: jan_c
"""

#gráficas con sales y fase sólida.

import pandas as pd
import matplotlib.pyplot as plt
import numpy as np 
from openpyxl import load_workbook

"con este codigo se llama a los datos"

from tkinter import *
from tkinter import filedialog


if __name__ == '__main__':
    
    def frame():
        
        def abrir_archivo():
            global archivo
            archivo = filedialog.askopenfilename(title="Abrir archivo .xlsx", initialdir="F:/", filetypes=(("Archivo .xlsx", "*.xlsx"), ("Archivo .xls", "*.xls")))
            raiz.destroy()
            
        raiz = Tk()
        mi_frame = Frame(raiz, width=200, height=60)
        mi_frame.pack()
        boton = Button(raiz, text="Abrir archivo", command=abrir_archivo)
        boton.pack(fill=X)
        boton.config(cursor="hand2")
        boton.config(bd=4)
        boton.config(relief="groove")
        raiz.mainloop()
        return archivo
    
    archivo = frame()
    data_0= pd.read_excel(archivo, "Espectro de fluorescencia 1_01", header=10, index_col=0)
    data_1= pd.read_excel(archivo, "Espectro de fluorescencia 2_02", header=10, index_col=0)
    data_2= pd.read_excel(archivo, "Espectro de fluorescencia 3_03", header=10, index_col=0)

    archivo_2 = frame()
    data_3= pd.read_excel(archivo_2, "datos_totales", header=0, index_col=0).round(1)
    #diff_cols = df2.columns.difference(df1.columns)
    indice_0 = pd.DataFrame(data_3.columns).values.astype(str) # convertimos los valores a texto
    
    #print(indice_0)
    # indices = indices[1:] # separamos la columna "Muestra"

    # Este ciclo separa el texto deseado usando indicadores entre corchetes
    def gen_indices(indice, i_inf, i_sup):
        nuevos_nombres = list()
        for i in indice:
            a = i[-1]
            d_r = (a[i_inf:i_sup])
            nuevos_nombres.append(d_r)
        return nuevos_nombres
    
    indice_1 = gen_indices(indice_0, 0, -2)
    indice_1 = [int(x) for x in indice_1]
     
    data_0 = data_0.loc[indice_1]
    data_1 = data_1.loc[indice_1]
    data_2 = data_2.loc[indice_1]
    
    indice_col = pd.DataFrame(data_0.columns).values.astype(str)
    indice_col = gen_indices(indice_col, 0, -6)
    
    indice_0 = gen_indices(indice_0, 0, 5)
    
    data_0 = data_0.set_axis(indice_col, axis = 1)
    data_1 = data_1.set_axis(indice_col, axis = 1)
    data_2 = data_2.set_axis(indice_col, axis = 1)
    data_0 = data_0.set_axis(indice_0, axis = 0)
    data_1 = data_1.set_axis(indice_0, axis = 0)
    data_2 = data_2.set_axis(indice_0, axis = 0)
    
    data_c = data_0.T
    data_c1 = data_1.T
    data_c2 = data_2.T
        
    
        
    reemplazo_de_texto = str(input("¿Quiere cambiar texto del indice?: "))
    
    if reemplazo_de_texto == "y":
         

        datos_complementarios = pd.concat([data_3, data_c, data_c1, data_c2], axis=0)
        datos_complementarios.index.name = 'Muestra'
        datos_complementarios.sort_index(inplace = True)
        datos_complementarios.round(1)
        print(datos_complementarios)
        
        indice_reemplazar = list(datos_complementarios.index)
        
        texto_receptor = str(input('¿cúal es el nombre correcto de receptor?: '))
        indice_c0 = [item.replace(indice_reemplazar[0], texto_receptor) for item in indice_reemplazar]
        datos_complementarios = datos_complementarios.reset_index()
        del(datos_complementarios["Muestra"])
        indice_correcto_1 = pd.DataFrame(indice_c0, columns=["Muestra"])
        datos_complementarios = pd.concat([indice_correcto_1, datos_complementarios], axis=1).set_index("Muestra")
        print(datos_complementarios)
   
        #archivo_3 = frame()
        #book = load_workbook(archivo_3)
        with pd.ExcelWriter(archivo_2[:-5] + "_final.xlsx") as writer:
            #writer.book = book 
            datos_complementarios.to_excel(writer, sheet_name = "datos_complementarios")
            
    else:
        
        datos_complementarios = pd.concat([data_3, data_c, data_c1, data_c2], axis=0)
        datos_complementarios.index.name = 'Muestra'
        datos_complementarios.sort_index(inplace = True)
        datos_complementarios.round(1)
    
        #archivo_3 = frame()
        #book = load_workbook(archivo_3)
        with pd.ExcelWriter(archivo_2[:-5] + "_final.xlsx") as writer:
            #writer.book = book 
            datos_complementarios.to_excel(writer, sheet_name = "datos_complementarios")
            